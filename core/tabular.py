"""Read a CSV or Excel file into (headers, rows) for batch runs.

``rows`` is a list of dicts keyed by the header names, with every value coerced
to a string (that's what gets typed). CSV needs no dependency; Excel uses
openpyxl (read-only, values not formulas).
"""

import csv
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

_EXCEL_EXTS = {".xlsx", ".xlsm"}


def read_table(path) -> tuple[list[str], list[dict]]:
    """Return (headers, rows). Rows that are entirely blank are skipped."""
    p = Path(path)
    if p.suffix.lower() in _EXCEL_EXTS:
        return _read_xlsx(p)
    return _read_csv(p)


def _cell(v) -> str:
    if v is None:
        return ""
    # openpyxl returns ints/floats/datetimes; keep integers clean (42 not 42.0).
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() if isinstance(v, str) else str(v)


def _dedupe_headers(raw: list[str]) -> list[str]:
    headers, seen = [], {}
    for i, h in enumerate(raw):
        name = (h or "").strip() or f"Column{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


def _read_csv(p: Path) -> tuple[list[str], list[dict]]:
    # utf-8-sig strips a BOM; sniff the delimiter so ';' and tab files work too.
    with open(p, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(f, dialect))

    if not rows:
        return [], []
    headers = _dedupe_headers(rows[0])
    data = []
    for r in rows[1:]:
        if not any((c or "").strip() for c in r):
            continue
        data.append({h: (r[i].strip() if i < len(r) and r[i] else "")
                     for i, h in enumerate(headers)})
    logger.debug("read_csv %s -> %d cols, %d rows", p.name, len(headers), len(data))
    return headers, data


def _read_xlsx(p: Path) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return [], []
        headers = _dedupe_headers([_cell(h) for h in header_row])
        data = []
        for row in it:
            if row is None or all(c is None or _cell(c) == "" for c in row):
                continue
            data.append({h: (_cell(row[i]) if i < len(row) else "")
                         for i, h in enumerate(headers)})
    finally:
        wb.close()
    logger.debug("read_xlsx %s -> %d cols, %d rows", p.name, len(headers), len(data))
    return headers, data
